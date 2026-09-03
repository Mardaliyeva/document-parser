"""Native XLSX adapter with formula-preserving worksheet extraction."""

from __future__ import annotations

import io
from collections import deque
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from typing import TYPE_CHECKING, Any, cast

from document_parser._adapter_utils import AssetCollector, normalize_text
from document_parser.exceptions import InvalidDocumentError, UnsafeDocumentError
from document_parser.models import (
    ContainerBlock,
    ContainerRole,
    ContentBlock,
    Diagnostic,
    DiagnosticSeverity,
    Document,
    DocumentFormat,
    DocumentMetadata,
    DocumentStatus,
    FigureBlock,
    SourceLocation,
    TableBlock,
    TableCell,
    TableRow,
    TextSpan,
)
from document_parser.results import AdapterOutput
from document_parser.sources import AdapterInput, ParseOptions

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


Coordinate = tuple[int, int]
Region = tuple[int, int, int, int]


class _XlsxContext:
    __slots__ = ("assets", "diagnostics", "partial", "sequence")

    def __init__(self, options: ParseOptions, source_name: str) -> None:
        self.assets = AssetCollector(options, source_name)
        self.diagnostics: list[Diagnostic] = []
        self.partial = False
        self.sequence = 0

    def next_id(self, kind: str) -> str:
        self.sequence += 1
        return f"xlsx:{kind}:{self.sequence:06d}"

    def warn(
        self,
        code: str,
        message: str,
        *,
        sheet_name: str | None = None,
        details: dict[str, object] | None = None,
        partial: bool = False,
    ) -> None:
        self.diagnostics.append(
            Diagnostic(
                code=code,
                message=message,
                severity=DiagnosticSeverity.WARNING,
                location=SourceLocation(sheet_name=sheet_name) if sheet_name else None,
                details=cast(dict[str, Any], details or {}),
            )
        )
        self.partial = self.partial or partial


class XlsxAdapter:
    """Convert worksheets, values, formulas, merges, and images into IR."""

    format = DocumentFormat.XLSX

    def parse(self, source: AdapterInput, options: ParseOptions) -> AdapterOutput:
        import defusedxml  # noqa: F401
        from openpyxl import load_workbook

        with source.open_binary() as stream:
            data = stream.read()
        try:
            formulas = load_workbook(
                io.BytesIO(data),
                read_only=False,
                data_only=False,
                keep_links=False,
                keep_vba=False,
            )
            cached = load_workbook(
                io.BytesIO(data),
                read_only=False,
                data_only=True,
                keep_links=False,
                keep_vba=False,
            )
        except Exception as exc:
            raise InvalidDocumentError(
                "XLSX workbook could not be opened", source_name=source.info.name
            ) from exc

        context = _XlsxContext(options, source.info.name)
        metadata = _workbook_metadata(formulas)
        sheets: list[ContentBlock] = []
        try:
            for sheet_index, worksheet in enumerate(formulas.worksheets):
                if worksheet.sheet_state != "visible" and not options.xlsx.include_hidden_sheets:
                    continue
                cached_sheet = cached[worksheet.title]
                blocks: list[ContentBlock] = list(
                    _worksheet_tables(worksheet, cached_sheet, context, options)
                )
                blocks.extend(_worksheet_images(worksheet, context))
                chart_count = len(cast(list[object], worksheet._charts))
                pivot_count = len(cast(list[object], worksheet._pivots))
                if chart_count or pivot_count:
                    context.warn(
                        "xlsx.drawing_omitted",
                        "Charts or pivot tables could not be rendered.",
                        sheet_name=worksheet.title,
                        details={"charts": chart_count, "pivots": pivot_count},
                        partial=True,
                    )
                block_id = context.next_id("sheet")
                sheets.append(
                    ContainerBlock(
                        block_id=block_id,
                        role=ContainerRole.SHEET,
                        title=(TextSpan(text=normalize_text(worksheet.title)),),
                        source=SourceLocation(
                            sheet_name=worksheet.title,
                            block_index=context.sequence,
                        ),
                        attributes={
                            "sheet_index": sheet_index,
                            "sheet_name": worksheet.title,
                            "visibility": worksheet.sheet_state,
                        },
                        blocks=tuple(blocks),
                    )
                )
        finally:
            formulas.close()
            cached.close()

        document = Document(
            document_id=f"sha256:{source.info.sha256}",
            source=source.info,
            metadata=metadata,
            blocks=tuple(sheets),
            assets=context.assets.refs,
            status=DocumentStatus.PARTIAL if context.partial else DocumentStatus.COMPLETE,
            diagnostics=tuple((*source.info.diagnostics, *context.diagnostics)),
        )
        return AdapterOutput(document=document, assets=context.assets.payloads)


def _workbook_metadata(workbook: Workbook) -> DocumentMetadata:
    properties = workbook.properties

    def aware(value: datetime | None) -> datetime | None:
        if value is None:
            return None
        return value.replace(tzinfo=UTC) if value.tzinfo is None else value

    author = normalize_text(properties.creator) if properties.creator else None
    keywords = tuple(
        item.strip()
        for item in normalize_text(properties.keywords or "").split(",")
        if item.strip()
    )
    return DocumentMetadata(
        title=normalize_text(properties.title) if properties.title else None,
        authors=(author,) if author else (),
        subject=normalize_text(properties.subject) if properties.subject else None,
        keywords=keywords,
        created_at=aware(properties.created),
        modified_at=aware(properties.modified),
        custom={"category": normalize_text(properties.category)} if properties.category else {},
    )


def _is_visible_cell(worksheet: Worksheet, row: int, column: int, options: ParseOptions) -> bool:
    from openpyxl.utils import get_column_letter

    if not options.xlsx.include_hidden_rows and worksheet.row_dimensions[row].hidden:
        return False
    letter = get_column_letter(column)
    return not (
        not options.xlsx.include_hidden_columns and worksheet.column_dimensions[letter].hidden
    )


def _occupied_cells(worksheet: Worksheet, options: ParseOptions) -> set[Coordinate]:
    stored = cast(dict[Coordinate, object], worksheet._cells)
    occupied = {
        coordinate
        for coordinate, cell in stored.items()
        if getattr(cell, "value", None) is not None
        and _is_visible_cell(worksheet, *coordinate, options)
    }
    for merged in worksheet.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                if _is_visible_cell(worksheet, row, column, options):
                    occupied.add((row, column))
    return occupied


def _connected_regions(coordinates: set[Coordinate]) -> tuple[Region, ...]:
    remaining = set(coordinates)
    result: list[Region] = []
    while remaining:
        start = min(remaining)
        queue = deque([start])
        remaining.remove(start)
        component = {start}
        while queue:
            row, column = queue.popleft()
            for neighbor in (
                (row - 1, column),
                (row + 1, column),
                (row, column - 1),
                (row, column + 1),
            ):
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    component.add(neighbor)
                    queue.append(neighbor)
        rows = [row for row, _ in component]
        columns = [column for _, column in component]
        result.append((min(rows), min(columns), max(rows), max(columns)))
    return tuple(sorted(result))


def _range_coordinates(region: Region) -> set[Coordinate]:
    min_row, min_column, max_row, max_column = region
    return {
        (row, column)
        for row in range(min_row, max_row + 1)
        for column in range(min_column, max_column + 1)
    }


def _formal_regions(worksheet: Worksheet) -> tuple[Region, ...]:
    from openpyxl.utils.cell import range_boundaries

    regions: list[Region] = []
    for table in sorted(worksheet.tables.values(), key=lambda item: item.ref):
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        regions.append((min_row, min_column, max_row, max_column))
    return tuple(regions)


def _worksheet_tables(
    worksheet: Worksheet,
    cached_sheet: Worksheet,
    context: _XlsxContext,
    options: ParseOptions,
) -> tuple[TableBlock, ...]:
    occupied = _occupied_cells(worksheet, options)
    formal = _formal_regions(worksheet)
    formal_coordinates: set[Coordinate] = set()
    for region in formal:
        formal_coordinates.update(_range_coordinates(region))
    remaining = occupied.difference(formal_coordinates)
    regions = (*formal, *_connected_regions(remaining))
    total_cells = sum(
        (max_row - min_row + 1) * (max_column - min_column + 1)
        for min_row, min_column, max_row, max_column in regions
    )
    if total_cells > options.xlsx.max_worksheet_cells:
        raise UnsafeDocumentError(
            "worksheet exceeds max_worksheet_cells", source_name=worksheet.title
        )
    return tuple(
        _region_table(worksheet, cached_sheet, region, index < len(formal), context, options)
        for index, region in enumerate(regions)
    )


def _scalar(value: object) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return normalize_text(value) if isinstance(value, str) else value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return normalize_text(str(value))


def _displayed(value: object) -> str:
    scalar = _scalar(value)
    if scalar is None:
        return ""
    if isinstance(scalar, bool):
        return "TRUE" if scalar else "FALSE"
    if isinstance(scalar, float):
        return format(scalar, ".15g")
    return str(scalar)


def _formula(value: object) -> str | None:
    if isinstance(value, str) and value.startswith("="):
        return normalize_text(value)
    text = str(value)
    return normalize_text(text) if text.startswith("=") else None


def _merged_origin(worksheet: Worksheet, row: int, column: int) -> tuple[int, int, int, int]:
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return merged.min_row, merged.min_col, merged.max_row, merged.max_col
    return row, column, row, column


def _region_table(
    worksheet: Worksheet,
    cached_sheet: Worksheet,
    region: Region,
    formal: bool,
    context: _XlsxContext,
    options: ParseOptions,
) -> TableBlock:
    from openpyxl.utils import get_column_letter

    min_row, min_column, max_row, max_column = region
    rows: list[TableRow] = []
    missing_formula_cache = 0
    for row in range(min_row, max_row + 1):
        cells: list[TableCell] = []
        for column in range(min_column, max_column + 1):
            if not _is_visible_cell(worksheet, row, column, options):
                continue
            origin_row, origin_column, merge_max_row, merge_max_column = _merged_origin(
                worksheet, row, column
            )
            if (row, column) != (origin_row, origin_column):
                continue
            cell = worksheet.cell(row=row, column=column)
            cached_cell = cached_sheet.cell(row=row, column=column)
            formula = _formula(cell.value)
            output_value = cached_cell.value if formula else cell.value
            if formula and output_value is None:
                missing_formula_cache += 1
                displayed = formula
            else:
                displayed = _displayed(output_value)
            cells.append(
                TableCell(
                    column_index=column - min_column,
                    row_span=merge_max_row - origin_row + 1,
                    column_span=merge_max_column - origin_column + 1,
                    is_header=formal and row == min_row,
                    raw_value=_scalar(cell.value),
                    displayed_text=displayed,
                    formula=formula,
                    number_format=normalize_text(str(cell.number_format)),
                )
            )
        rows.append(TableRow(row_index=row - min_row, cells=tuple(cells)))

    start = f"{get_column_letter(min_column)}{min_row}"
    end = f"{get_column_letter(max_column)}{max_row}"
    cell_range = start if start == end else f"{start}:{end}"
    if missing_formula_cache:
        context.warn(
            "xlsx.formula_cache_missing",
            "Some formulas did not contain cached display values.",
            sheet_name=worksheet.title,
            details={"cell_range": cell_range, "count": missing_formula_cache},
        )
    block_id = context.next_id("table")
    hidden_rows = [
        row for row in range(min_row, max_row + 1) if worksheet.row_dimensions[row].hidden
    ]
    hidden_columns = [
        get_column_letter(column)
        for column in range(min_column, max_column + 1)
        if worksheet.column_dimensions[get_column_letter(column)].hidden
    ]
    return TableBlock(
        block_id=block_id,
        row_count=max_row - min_row + 1,
        column_count=max_column - min_column + 1,
        rows=tuple(rows),
        source=SourceLocation(
            sheet_name=worksheet.title,
            cell_range=cell_range,
            block_index=context.sequence,
        ),
        attributes=cast(
            dict[str, Any],
            {
                "formal_table": formal,
                "hidden_rows": hidden_rows,
                "hidden_columns": hidden_columns,
            },
        ),
    )


def _worksheet_images(worksheet: Worksheet, context: _XlsxContext) -> tuple[FigureBlock, ...]:
    from openpyxl.utils import get_column_letter

    figures: list[FigureBlock] = []
    for image_index, image in enumerate(cast(list[Any], worksheet._images)):
        try:
            data = bytes(image._data())
            filename = str(getattr(image, "path", "") or f"image-{image_index + 1}.bin")
            image_format = str(getattr(image, "format", "") or "").lower()
            media_type = f"image/{image_format}" if image_format else None
            ref = context.assets.add(data, filename=filename, media_type=media_type)
        except UnsafeDocumentError:
            raise
        except Exception:
            context.warn(
                "xlsx.image_omitted",
                "A worksheet image could not be extracted.",
                sheet_name=worksheet.title,
                partial=True,
            )
            continue
        anchor = image.anchor
        cell_range: str | None = anchor if isinstance(anchor, str) else None
        if cell_range is None and hasattr(anchor, "_from"):
            marker = anchor._from
            cell_range = f"{get_column_letter(int(marker.col) + 1)}{int(marker.row) + 1}"
        block_id = context.next_id("figure")
        figures.append(
            FigureBlock(
                block_id=block_id,
                asset_id=ref.asset_id,
                source=SourceLocation(
                    sheet_name=worksheet.title,
                    cell_range=cell_range,
                    block_index=context.sequence,
                    asset_id=ref.asset_id,
                ),
                attributes={"image_index": image_index},
            )
        )
    return tuple(figures)
